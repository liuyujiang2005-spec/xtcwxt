"""
物流财务系统 - 表格解析服务
==========================
两种使用方式：

方式一（独立工具）：
  python3 table_parser.py --file 账单.xlsx
  → 输出解析后的JSON文件

方式二（API服务，给系统调用）：
  python3 table_parser.py --serve
  → 启动HTTP服务，系统通过API上传文件解析

安装依赖：
  pip install openpyxl openai fastapi uvicorn python-multipart

环境变量（方式二用）：
  OPENAI_API_KEY=***  # GPT-4o API Key
  DEEPSEEK_API_KEY=***  # DeepSeek API Key（可选，用于分类）
"""

import json, os, sys, io, argparse, base64
from pathlib import Path
from typing import Optional
from openai import OpenAI

# ========== 配置 ==========
# 方式一：直接修改这里
OPENAI_KEY = os.environ.get("OPENAI_API_KEY", "")
DEEPSEEK_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
DEEPSEEK_MODEL = "deepseek-v4-flash"
# ==========================

# ──────────────────────────────────────────
# 核心：GPT-4o 分析表格结构
# ──────────────────────────────────────────

def analyze_structure(ws, filename=""):
    """让GPT-4o分析表格结构，返回解析规则"""
    merged = list(ws.merged_cells.ranges)
    merged_str = ", ".join([str(m) for m in merged[:15]]) if merged else "无"
    
    # 取样本数据（前30行 + 最后5行）
    sample = []
    for ri in range(1, min(ws.max_row + 1, 35)):
        cells = []
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                cells.append(f"C{ci}={str(v).strip()[:45]}")
        if cells:
            sample.append(f"R{ri}: {' | '.join(cells)}")
    
    if ws.max_row > 35:
        for ri in range(ws.max_row - 4, ws.max_row + 1):
            cells = []
            for ci in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    cells.append(f"C{ci}={str(v).strip()[:45]}")
            if cells:
                sample.append(f"R{ri}: {' | '.join(cells)}")
    
    # 找列头行：扫描前10行找包含"日期""金额""编号"等关键词的行
    header_candidates = ["日期", "编号", "金额", "名称", "数量", "单价", "总价", "备注", "状态"]
    header_row = 1
    for ri in range(1, min(11, ws.max_row + 1)):
        row_text = ""
        for ci in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(row=ri, column=ci).value
            if v:
                row_text += str(v)
        match_count = sum(1 for kw in header_candidates if kw in row_text)
        if match_count >= 3:
            header_row = ri
            break
    
    # 获取列头
    headers = []
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=ci).value
        headers.append(str(v).strip() if v else f"列{ci}")
    
    prompt = f"""分析这个装柜清单Excel表格的结构，返回JSON格式的解析规则。

文件名：{filename}
规格：{ws.max_row}行 x {ws.max_column}列
合并单元格：{len(merged)}个
示例合并：{merged_str}

列头（推测在第{header_row}行）：
{' | '.join(headers)}

样本数据（前30行 + 最后5行）：
{chr(10).join(sample)}

这是一个固定的22列装柜单格式，列顺序为：
日期、唛头、仓库、运输方式、运单号、货型、品名、尺寸、件数、国内单号、单项体积、单项重量、总体积、总重量、计费体积、总计费体积、单价、单项价格、订单总价、备注、结算状态、柜号

其中9个字段是合并单元格（头层，一个运单号对应一条）：
日期、唛头、仓库、运输方式、运单号、总体积、总重量、总计费体积、订单总价
合并单元格只在第一行有值，后续行为空，需要用运单号做分组标识进行前向填充。

13个字段是逐行明细：
货型、品名、尺寸（为字符串格式如"20×30×40"）、件数、国内单号、单项体积、单项重量、计费体积、单价、单项价格、备注、结算状态、柜号

规则（只返回JSON，不要其他文字，JSON格式如下）：
{{
  "title_rows": 标题行数（数字，列头之前的行数）,
  "header_row": 列头所在行号（数字）,
  "data_start_row": 数据起始行号（数字）,
  "order_id_column": "运单号",
  "date_column": "日期",
  "customer_column": "唛头",
  "total_price_column": "订单总价",
  "has_merge_cells": true,
  "merge_columns": ["日期", "唛头", "仓库", "运输方式", "运单号", "总体积", "总重量", "总计费体积", "订单总价"],
  "merge_rule": "以运单号分组，头层字段只在该组第一行有值，后续行向前填充",
  "product_detail_columns": ["货型", "品名", "尺寸", "件数", "国内单号", "单项体积", "单项重量", "计费体积", "单价", "单项价格", "备注", "结算状态", "柜号"],
  "parse_logic": "运单号相同的行属于同一订单，头层字段前向填充，明细字段逐行解析"
}}
"""
    
    client = OpenAI(api_key=OPENAI_KEY)
    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2000,
        response_format={"type": "json_object"}
    )
    
    return json.loads(resp.choices[0].message.content)


# ──────────────────────────────────────────
# 核心：解析数据
# ──────────────────────────────────────────

def parse_data(ws, rules):
    """根据GPT-4o规则解析表格"""
    data_start = rules.get("data_start_row", rules.get("header_row", 7) + 1)
    header_row = rules.get("header_row", 7)
    order_col_name = rules.get("order_id_column", "汇诚邦单号")

    # 校验 header_row：检查是否包含多个已知列头关键词
    HEADER_KEYWORDS = ["日期", "唛头", "仓库", "运输", "运单", "货型", "品名", "尺寸", "件数",
                       "国内单号", "单项体积", "单项重量", "总体积", "总重量", "计费体积", "单价", "单项价格", "订单总价", "备注", "结算", "柜号"]
    
    def count_header_matches(row_idx):
        text = ""
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=ci).value
            if v: text += str(v)
        return sum(1 for kw in HEADER_KEYWORDS if kw in text)

    match_count = count_header_matches(header_row)
    if match_count < 3:
        # GPT-4o 误判了 header_row，扫描前10行找最佳列头行
        best_row, best_cnt = 1, 0
        for ri in range(1, min(ws.max_row + 1, 11)):
            cnt = count_header_matches(ri)
            if cnt > best_cnt:
                best_cnt, best_row = cnt, ri
        if best_cnt >= 3:
            header_row = best_row
    
    # 建立列名→列索引映射
    col_map = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=ci).value
        if v:
            col_map[str(v).strip()] = ci
    
    # 找订单标识列：让GPT-4o选，但用代码验证
    order_col = None
    order_col_name_used = order_col_name
    
    # 候选列：GPT-4o推荐的 + 常见订单号列名
    candidates = [order_col_name]
    if order_col_name not in ["单号", "汇诚邦单号", "订单号", "运单号", "编号"]:
        candidates += ["单号", "汇诚邦单号", "订单号", "运单号", "编号", "柜号", "提单号"]
    
    # 尝试每个候选列，看哪个分组最合理
    best_col = None
    best_score = -1
    
    for candidate in candidates:
        col_idx = None
        for name, idx in col_map.items():
            if candidate in name:
                col_idx = idx
                break
        if not col_idx:
            continue
        
        # 统计该列有值的行数和非重复值数
        non_empty = 0
        unique_vals = set()
        for ri in range(data_start, ws.max_row + 1):
            v = ws.cell(row=ri, column=col_idx).value
            if v is not None and str(v).strip():
                non_empty += 1
                unique_vals.add(str(v).strip()[:50])
        
        # 评分：非重复值数量适中（太多=每行一个单号，太少=全部一样）
        total_rows = ws.max_row - data_start + 1
        if non_empty == 0:
            continue
        ratio = len(unique_vals) / non_empty if non_empty > 0 else 0
        
        # 理想情况：30-50%的行有值（每2-3行一个订单），非重复值占80%以上
        if 0.2 <= ratio <= 1.0 and len(unique_vals) >= 2:
            score = (1 - abs(ratio - 0.4)) * 100 + min(len(unique_vals), 100)
            if score > best_score:
                best_score = score
                best_col = col_idx
                best_name = candidate
                best_unique = len(unique_vals)
    
    # 用评分最高的列
    if best_col:
        order_col = best_col
        order_col_name_used = best_name
    
    if not order_col:
        order_col = 5
        if len(col_map) >= 5:
            order_col = list(col_map.values())[4]
            order_col_name_used = list(col_map.keys())[4]

    # 修正 data_start：GPT-4o 可能因合并单元格误判，向上扫描找第一条有运单号的行
    for ri in range(data_start - 1, header_row, -1):
        v = ws.cell(row=ri, column=order_col).value
        if v is not None and str(v).strip():
            data_start = ri
            break

    # 逐行解析
    orders = []
    current_order = None
    
    for ri in range(data_start, ws.max_row + 1):
        row_data = {}
        for name, ci in col_map.items():
            row_data[name] = ws.cell(row=ri, column=ci).value
        
        order_val = ws.cell(row=ri, column=order_col).value
        order_str = str(order_val).strip() if order_val is not None else ""
        
        # 判断是否新订单：order列有值即为新订单开始
        is_new = bool(order_str)
        
        if is_new:
            if current_order:
                orders.append(current_order)
            current_order = {"订单号": order_str, "产品明细": []}
            for name in col_map:
                if row_data[name] is not None:
                    current_order[name] = row_data[name]
            current_order["产品明细"].append(dict(row_data))
        else:
            if current_order is not None:
                # 前向填充：合并单元格字段用上一行值
                for name in col_map:
                    if row_data[name] is None and name in current_order:
                        row_data[name] = current_order[name]
                if row_data.get('品名') or row_data.get('尺寸') or row_data.get('件数'):
                    current_order["产品明细"].append(dict(row_data))
    
    if current_order:
        orders.append(current_order)
    
    return orders


# ──────────────────────────────────────────
# 核心：DeepSeek 分类
# ──────────────────────────────────────────

def deepseek_classify(orders):
    """用DeepSeek对解析结果做分类"""
    if not DEEPSEEK_KEY:
        return None
    
    client = OpenAI(
        api_key=DEEPSEEK_KEY,
        base_url="https://api.deepseek.com/v1"
    )
    
    summary = {
        "总订单数": len(orders),
        "总产品数": sum(len(o.get("产品明细", [])) for o in orders),
        "总金额": sum(float(o.get(list(o.keys())[-2], 0) or 0) for o in orders if isinstance(o.get(list(o.keys())[-2]), (int, float, str))),
    }
    
    prompt = f"""分析以下物流账单数据，按客户、运输方式、货物类型进行分类汇总。

总览：共{summary['总订单数']}笔订单，{summary['总产品数']}个产品

前10笔订单详情：
{json.dumps(orders[:10], ensure_ascii=False, indent=2)}

请输出JSON格式的分类结果：
{{
  "按客户汇总": [{{"客户": "名称", "订单数": 数字, "产品数": 数字, "总金额": 数字}}],
  "货物类型分布": [{{"类型": "名称", "数量": 数字}}],
  "异常数据": [{{"问题": "说明", "订单号": "编号"}}]
}}
"""
    
    resp = client.chat.completions.create(
        model=DEEPSEEK_MODEL,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2000,
        response_format={"type": "json_object"}
    )
    
    return json.loads(resp.choices[0].message.content)


# ──────────────────────────────────────────
# 方式一：CLI工具
# ──────────────────────────────────────────

def run_cli(file_path: str, output_path: Optional[str] = None):
    """命令行模式"""
    import openpyxl
    
    print(f"📂 读取: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    print(f"   工作表: {ws.title}")
    print(f"   大小: {ws.max_row}行 x {ws.max_column}列")
    print(f"   合并单元格: {len(list(ws.merged_cells.ranges))}个")
    
    print("\n🔍 Step 1: GPT-4o分析结构...")
    rules = analyze_structure(ws, file_path)
    print(f"   数据起始行: {rules.get('data_start_row')}")
    print(f"   订单标识列: {rules.get('order_id_column')}")
    print(f"   解析规则: {rules.get('parse_logic', '')[:80]}...")
    
    print("\n📊 Step 2: 解析数据...")
    orders = parse_data(ws, rules)
    print(f"   共 {len(orders)} 笔订单")
    
    print("\n📋 前3笔:")
    for i, order in enumerate(orders[:3]):
        print(f"\n--- 订单 {i+1} ---")
        for k, v in order.items():
            if k != "产品明细":
                print(f"   {k}: {v}")
        prods = order.get("产品明细", [])
        print(f"   产品数: {len(prods)}")
        for j, p in enumerate(prods[:3]):
            pname = p.get("品名", p.get("商品名称", p.get("产品名称", "?")))
            qty = p.get("件数", p.get("数量", "?"))
            price = p.get("单项价格", p.get("金额", "?"))
            print(f"     产品{j+1}: {pname} x{qty} ¥{price}")
        if len(prods) > 3:
            print(f"     ...还有 {len(prods)-3} 个产品")
    
    print(f"\n🧠 DeepSeek 分类分析...")
    classify = deepseek_classify(orders)
    if classify:
        print(f"   按客户: {len(classify.get('按客户汇总', []))}个客户")
        if classify.get("异常数据"):
            print(f"   ⚠️ 发现 {len(classify['异常数据'])} 个异常")
            for a in classify["异常数据"]:
                print(f"     {a}")
    
    # 保存
    if not output_path:
        output_path = file_path.rsplit(".", 1)[0] + "_解析结果.json"
    
    result = {
        "文件名": file_path,
        "规则": rules,
        "订单数": len(orders),
        "订单": orders,
        "分类汇总": classify
    }
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n💾 保存: {output_path}")
    return orders


# ──────────────────────────────────────────
# 方式二：API服务
# ──────────────────────────────────────────

def run_api_server(host: str = "0.0.0.0", port: int = 8800):
    """启动FastAPI服务"""
    try:
        from fastapi import FastAPI, UploadFile, File, Form, HTTPException
        from fastapi.middleware.cors import CORSMiddleware
        import uvicorn, tempfile, openpyxl
    except ImportError:
        print("请安装: pip install fastapi uvicorn python-multipart")
        sys.exit(1)
    
    app = FastAPI(title="表格解析服务", version="1.0")
    app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
    
    @app.post("/api/table/parse")
    async def parse_table(file: UploadFile = File(...), classify: bool = Form(False)):
        """上传Excel，返回解析结果"""
        if not file.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(400, "仅支持 .xlsx / .xls 文件")
        
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        rules = analyze_structure(ws, file.filename)
        orders = parse_data(ws, rules)
        
        result = {
            "文件名": file.filename,
            "规则": rules,
            "订单数": len(orders),
            "订单": orders,
            "分类汇总": None
        }
        
        if classify:
            result["分类汇总"] = deepseek_classify(orders)
        
        return result
    
    @app.post("/api/table/analyze")
    async def analyze_only(file: UploadFile = File(...)):
        """只分析结构，不解析全表"""
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rules = analyze_structure(ws, file.filename)
        return {"文件名": file.filename, "规则": rules}
    
    @app.get("/api/health")
    def health():
        return {
            "status": "ok",
            "gpt4o": "已配置" if OPENAI_KEY else "未配置",
            "deepseek": "已配置" if DEEPSEEK_KEY else "未配置"
        }
    
    print(f"\n🚀 API服务启动: http://{host}:{port}")
    print(f"   POST /api/table/parse    上传Excel解析")
    print(f"   POST /api/table/analyze  仅分析结构")
    print(f"   GET  /api/health         健康检查")
    print(f"\n在系统里调这些接口就行")
    uvicorn.run(app, host=host, port=port)


# ──────────────────────────────────────────
# 入口
# ──────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="物流财务系统中表格解析")
    parser.add_argument("--file", "-f", help="Excel文件路径")
    parser.add_argument("--output", "-o", help="输出JSON路径（可选）")
    parser.add_argument("--serve", "-s", action="store_true", help="启动API服务")
    parser.add_argument("--port", type=int, default=8800, help="API服务端口")
    parser.add_argument("--host", default="0.0.0.0", help="API服务地址")
    args = parser.parse_args()
    
    if args.serve:
        run_api_server(args.host, args.port)
    elif args.file:
        run_cli(args.file, args.output)
    else:
        parser.print_help()
        print("\n示例：")
        print("  解析文件:  python3 table_parser.py --file 账单.xlsx")
        print("  启动服务:  python3 table_parser.py --serve --port 8800")
