import openpyxl
wb = openpyxl.load_workbook('/root/xtcwxt/public/bill_template.xlsx')
ws = wb.worksheets[0]
print("Row 5 headers:")
for cell in ws[5]:
    if cell.value: print(f"  Col{cell.column}: {cell.value}")
print(f"Max col before: {ws.max_column}")
# Check what's in col 22
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 22).value
    if v: print(f"  R{r}C22: {v}")
ws.delete_cols(22, 1)
wb.save('/root/xtcwxt/public/bill_template.xlsx')
print(f"Max col after: {ws.max_column}")
print("Done")
